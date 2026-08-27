from pathlib import Path
from database import VoterDatabase
from excel_writer import ExcelWriter
from matcher import Matcher
import openpyxl

excel_path = Path(r'C:\Users\chenn\Downloads\59_EF_Filled_Forms_Complete_1-8-2026.xlsx')
db = VoterDatabase()

writer = ExcelWriter(str(excel_path))
writer.load()
epic_rows = writer.read_epic_rows()
matcher = Matcher(db)
results, stats = matcher.match_epics(epic_rows)
print('Total matched', len([r for r in results if r.found]))

wb = openpyxl.Workbook()
ws = wb.active

for res in results:
    if not res.found:
        continue
    try:
        ws.cell(row=res.row_index, column=1).value = res.house_no
        ws.cell(row=res.row_index, column=2).value = res.colony
    except Exception as e:
        print('ERROR at row', res.row_index, 'EPIC', res.epic)
        print('house_no repr', repr(res.house_no))
        print('colony repr', repr(res.colony))
        print('type', type(e).__name__, e)
        raise
print('No illegal chars found')
