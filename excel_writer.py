"""
excel_writer.py
================
Reads the user-supplied Excel template, hands EPIC numbers to
matcher.py, then writes House No / Colony back into the SAME
workbook object (loaded with openpyxl, never re-created from scratch)
so existing formatting, colors, and formulas are preserved untouched.
Saves the result as Completed.xlsx in the chosen output folder.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

import config
from matcher import MatchResult, MatchStats, Matcher
from database import VoterDatabase
from utils import logger, safe_str


class ExcelDataError(RuntimeError):
    """Raised when matched EPIC rows have no usable extracted House No / Colony."""

_HEADER_NAME_MAP = {
    "epic number": config.COL_EPIC,
    "epic no": config.COL_EPIC,
    "epic": config.COL_EPIC,
    "epic #": config.COL_EPIC,
    "epic no.": config.COL_EPIC,
    "h no": config.COL_HOUSE_NO,
    "house no": config.COL_HOUSE_NO,
    "house number": config.COL_HOUSE_NO,
    "h no flat no": config.COL_HOUSE_NO,
    "h no / flat no": config.COL_HOUSE_NO,
    "h.no flat no": config.COL_HOUSE_NO,
    "flat no": config.COL_HOUSE_NO,
    "flat number": config.COL_HOUSE_NO,
    "house": config.COL_HOUSE_NO,
    "flat": config.COL_HOUSE_NO,
    "colony": config.COL_COLONY,
    "area": config.COL_COLONY,
    "locality": config.COL_COLONY,
    "area colony": config.COL_COLONY,
    "area / colony": config.COL_COLONY,
    "colony area": config.COL_COLONY,
    "area / locality": config.COL_COLONY,
    "address": config.COL_ADDRESS,
    "full address": config.COL_ADDRESS,
}


def _normalize_header_name(name: str) -> str:
    return " ".join(re.sub(r"[^\w]+", " ", safe_str(name)).split()).lower()


def _map_header_name(name: str) -> Optional[str]:
    return _HEADER_NAME_MAP.get(_normalize_header_name(name))


class ExcelColumnError(RuntimeError):
    """Raised when required columns are missing from the template."""


class ExcelWriter:
    def __init__(self, excel_path: str) -> None:
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet: Optional[Worksheet] = None
        self.header_map: dict[str, int] = {}
        self.header_row = 1
        self.raw_header_names: List[str] = []

    # -- loading / validation --------------------------------------------------
    def load(self) -> None:
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=False)
        except FileNotFoundError as exc:
            raise ExcelColumnError(f"Excel file not found: {self.excel_path}") from exc
        except PermissionError as exc:
            raise ExcelColumnError(
                f"Permission denied opening Excel file (is it open in another "
                f"program?): {self.excel_path}"
            ) from exc
        except Exception as exc:
            raise ExcelColumnError(f"Could not open Excel file (corrupt?): {exc}") from exc

        self.worksheet = self.workbook.active
        self._build_header_map()
        self._validate_required_columns()

    def _build_header_map(self) -> None:
        self.header_map.clear()
        self.raw_header_names = []
        self.header_row = self._find_header_row()
        for cell in self.worksheet[self.header_row]:
            if cell.value is None:
                continue
            raw_header = safe_str(cell.value).strip()
            self.raw_header_names.append(raw_header)
            canonical = _map_header_name(raw_header)
            if canonical:
                self.header_map[canonical] = cell.column

    def _find_header_row(self) -> int:
        candidate_rows: list[tuple[int, int, list[str]]] = []
        max_search_row = min(self.worksheet.max_row, 20)

        for row_idx in range(1, max_search_row + 1):
            row_matches = set()
            row_cells: list[str] = []
            for cell in self.worksheet[row_idx]:
                if cell.value is None:
                    continue
                raw_header = safe_str(cell.value).strip()
                row_cells.append(raw_header)
                canonical = _map_header_name(raw_header)
                if canonical:
                    row_matches.add(canonical)
            if row_matches:
                candidate_rows.append((len(row_matches), row_idx, row_cells))

        if candidate_rows:
            _, header_row, _ = max(candidate_rows, key=lambda item: (item[0], -item[1]))
            return header_row

        return 1

    def _col(self, name: str) -> Optional[int]:
        return self.header_map.get(name.strip())

    def _validate_required_columns(self) -> None:
        missing = [c for c in config.REQUIRED_EXCEL_COLUMNS if self._col(c) is None]
        if missing:
            found_columns = self.raw_header_names or list(self.header_map.keys())
            raise ExcelColumnError(
                f"Excel template is missing required column(s): {', '.join(missing)}. "
                f"Found columns: {found_columns}"
            )

    # -- reading -----------------------------------------------------------------
    def read_epic_rows(self) -> List[tuple[int, str]]:
        """Returns (row_index, epic_value) for every data row below the header."""
        epic_col = self._col(config.COL_EPIC)
        rows: List[tuple[int, str]] = []
        for row_idx in range(self.header_row + 1, self.worksheet.max_row + 1):
            cell = self.worksheet.cell(row=row_idx, column=epic_col)
            if cell.value is None and self._row_is_blank(row_idx):
                continue
            rows.append((row_idx, safe_str(cell.value)))
        return rows

    def _row_is_blank(self, row_idx: int) -> bool:
        return all(
            self.worksheet.cell(row=row_idx, column=c).value in (None, "")
            for c in range(1, self.worksheet.max_column + 1)
        )

    # -- writing -------------------------------------------------------------------
    def apply_results(self, results: List[MatchResult]) -> None:
        house_col = self._col(config.COL_HOUSE_NO)
        colony_col = self._col(config.COL_COLONY)
        address_col = self._col(config.COL_ADDRESS)

        if address_col is None:
            address_col = self._add_address_column()

        for res in results:
            if res.found:
                if house_col:
                    self.worksheet.cell(
                        row=res.row_index,
                        column=house_col,
                    ).value = safe_str(res.house_no)
                if colony_col:
                    self.worksheet.cell(
                        row=res.row_index,
                        column=colony_col,
                    ).value = safe_str(res.colony)
                if address_col:
                    self.worksheet.cell(
                        row=res.row_index,
                        column=address_col,
                    ).value = safe_str(res.address)
            else:
                if house_col:
                    self.worksheet.cell(row=res.row_index, column=house_col).value = None
                if colony_col:
                    self.worksheet.cell(row=res.row_index, column=colony_col).value = None
                if address_col:
                    self.worksheet.cell(row=res.row_index, column=address_col).value = None

    def _add_address_column(self) -> Optional[int]:
        """If the Address column is missing, append it to the header row."""
        if self.worksheet is None:
            return None

        target_col = self.worksheet.max_column + 1
        self.worksheet.cell(row=self.header_row, column=target_col).value = config.COL_ADDRESS
        self.header_map[config.COL_ADDRESS] = target_col
        return target_col

    def save(self, output_folder: str, filename: str = "Completed.xlsx") -> str:
        out_path = Path(output_folder) / filename
        try:
            self.workbook.save(out_path)
        except PermissionError as exc:
            raise ExcelColumnError(
                f"Permission denied saving to {out_path}. Close the file if it's "
                f"open elsewhere, or choose a different output folder."
            ) from exc
        logger.info("Saved completed workbook to %s", out_path)
        return str(out_path)


def run_matching_pipeline(
    excel_path: str,
    db: VoterDatabase,
    output_folder: str,
    output_filename: str = "Completed.xlsx",
    progress_callback=None,
) -> tuple[str, MatchStats]:
    """High-level convenience function used by the GUI: load Excel,
    match every EPIC, write results back, save, return (output_path, stats)."""
    writer = ExcelWriter(excel_path)
    writer.load()

    epic_rows = writer.read_epic_rows()
    matcher = Matcher(db)
    results, stats = matcher.match_epics(epic_rows, progress_callback=progress_callback)

    matched_results = [result for result in results if result.found]
    if matched_results:
        # Print first 20 extracted matched records to stdout (user-requested)
        to_print = matched_results[:20]
        print("First {} matched records:".format(min(20, len(matched_results))))
        for result in to_print:
            print()
            print("EPIC : {}".format(result.epic or "<blank>"))
            print("House No : {}".format(result.house_no or ""))
            print("Colony : {}".format(result.colony or ""))
            print("Address : {}".format(result.address or ""))

        # If any matched result is missing house_no or colony, abort for debugging
        for result in matched_results:
            if not result.house_no or not result.colony:
                logger.error(
                    "Matched EPIC %s (row %s) missing house_no or colony: house_no=%r colony=%r",
                    result.epic,
                    result.row_index,
                    result.house_no,
                    result.colony,
                )
                raise ExcelDataError(
                    f"Matched EPIC {result.epic} (row {result.row_index}) missing House No or Colony. Aborting to debug parser."
                )
    else:
        logger.info(
            "No matched EPIC records were found in the Excel file. House No and Colony will remain blank."
        )

    writer.apply_results(results)
    out_path = writer.save(output_folder, filename=output_filename)
    return out_path, stats
